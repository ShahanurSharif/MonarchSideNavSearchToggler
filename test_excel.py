#!/usr/bin/env python3
"""
Simple test script to check Excel file content
"""

import json
import sys
import os

def test_excel_structure():
    """Test Excel file structure without pandas"""
    excel_file = "menu_items.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file '{excel_file}' not found")
        return
    
    print(f"✅ Excel file found: {excel_file}")
    print(f"📦 File size: {os.path.getsize(excel_file)} bytes")
    
    # Try to read with openpyxl if available
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print(f"📋 Sheet name: {ws.title}")
        print(f"📊 Rows: {ws.max_row}")
        print(f"📊 Columns: {ws.max_column}")
        
        # Read first few rows
        print("\n📋 First 5 rows:")
        for row in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else '')
            print(f"  Row {row}: {row_data}")
            
    except ImportError:
        print("❌ openpyxl not available. Install with: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")

def test_config_structure():
    """Test config file structure"""
    config_file = "sharepoint/assets/monarchSidebarNavConfig.json"
    
    if os.path.exists(config_file):
        print(f"\n✅ Config file found: {config_file}")
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            print(f"📋 Config keys: {list(config.keys())}")
            print(f"📊 Items count: {len(config.get('items', []))}")
            
            if config.get('items'):
                print(f"📋 Sample items: {config['items'][:2]}")
            else:
                print("⚠️  No items in config")
                
        except Exception as e:
            print(f"❌ Error reading config: {e}")
    else:
        print(f"\n❌ Config file not found: {config_file}")

if __name__ == "__main__":
    print("🧪 Testing Excel and Config files...")
    test_excel_structure()
    test_config_structure() 